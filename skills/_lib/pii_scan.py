#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""XLSX/文字列の簡易 PII プリフライト（v3.2.0〜）。

`/template-create --scope global` と `/template-promote` の前に、テンプレート
XLSX に **実クライアントのデータが残っていないか** を決定論的にスキャンする。

これは置換や削除は行わない。検出結果（カテゴリ・セル位置・抜粋）を返すのみ。
最終判断は必ず弁護士に委ねる（false positive が多い領域のため）。

## 検出カテゴリ
- `postal_code`        — 〒100-0001 形式の郵便番号
- `phone`              — 03-1234-5678 / 090-1234-5678 形式の電話番号
- `bank_account`       — 「口座番号 1234567」「普通預金 1234567」等の連続 7 桁数字
- `birthdate`          — 「生年月日」「S.50.1.1」「昭和50年1月1日」等
- `personal_name`      — 「○○様」「○○氏」「原告 ○○」等
- `address_jp`         — 都道府県名 + 市区町村 を含む住所形式
- `my_number`          — 個人番号らしき 12 桁連続数字（文脈なしでも検出）
- `case_number`        — 「令和○年(ワ)第○号」等の実在しそうな事件番号

各カテゴリは conservative（偽陽性寄り）に設計。false negative より false
positive を優先する（PII 混入見逃しのほうが深刻なため）。

## CLI
```
python3 skills/_lib/pii_scan.py scan --xlsx <path> [--sheet <name>] [--json]
python3 skills/_lib/pii_scan.py check <text>   # テキスト 1 文を判定（テスト用）
```
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# 検出パターン
# ---------------------------------------------------------------------------

# 全角→半角数字変換用（phone/postal のマッチで使う）
_FULLWIDTH_DIGIT_RE = re.compile(r"[０-９]")

def _normalize_digits(text: str) -> str:
    """全角数字・全角ハイフンを半角に直した文字列を返す（マッチ判定用）。"""
    return text.translate(str.maketrans("０１２３４５６７８９－ー−", "0123456789---"))

# 「〒100-0001」「〒１００−０００１」（全角数字・全角ハイフン・Unicode minus 対応）
_POSTAL_RE = re.compile(r"〒\s*[\d０-９]{3}[\-－‐‑ー−]?[\d０-９]{4}")

# 電話/FAX: 03-1234-5678, 090-1234-5678, (03)1234-5678, 全角 ０３−１２３４−５６７８
# 区切り文字: ASCII ハイフン, 全角ハイフン (－), Unicode minus (−), ソフトハイフン (‐),
# 長音 (ー), カッコ類
_PHONE_SEP = r"[\-－−‐‑ー(（)）]"
_PHONE_RE = re.compile(
    r"(?:\bTEL|\bFAX|電話|ＴＥＬ|ＦＡＸ|Tel|Fax)?\s*[:：]?\s*"
    r"(?:"
    r"[0０][\d０-９]{1,4}" + _PHONE_SEP + r"[\d０-９]{1,4}" + _PHONE_SEP + r"[\d０-９]{3,4}"
    r"|\b[0０][\d０-９]{9,10}\b"
    r")"
)

# 銀行口座: 「口座番号 1234567」「普通預金 1234567」「当座」「店番号」「支店番号」
# 「口座名義: 甲野太郎」も拾う（氏名ラベル付き。ラベルなしは personal_name に任せる）
_BANK_ACCOUNT_RE = re.compile(
    r"(?:口座番号|普通預金|当座預金|貯蓄預金|店番号?|支店番号|Account)\s*[:：\-]?\s*[\d０-９]{7,}"
)
_BANK_HOLDER_RE = re.compile(
    r"(?:口座名義|ご?名義人?|名義人)\s*[:：\-]?\s*[一-龥々ヶァ-ヺー・\s]{2,20}"
)

# 生年月日
_BIRTHDATE_RE = re.compile(
    r"(?:生年月日|生年|誕生日|ご?年齢|DOB)\s*[:：]|"
    r"(?:昭和|平成|令和|大正|明治)\s*(?:元|[\d０-９]{1,2})\s*年\s*[\d０-９]{1,2}\s*月\s*[\d０-９]{1,2}\s*日|"
    r"\b[SHRTM]\.?\s*\d{1,2}\.?\d{1,2}\.?\d{1,2}\b"
)

# 当事者氏名: 原告/被告/弁護士/代理人 + 氏名
# 中黒（・）を含む外国人氏名（ジョン・スミス）も検出するため [一-龥々ヶァ-ヺー・] を採用
_NAME_RE = re.compile(
    r"(?:原告|被告|債権者|債務者|申立人|相手方|弁護士|代理人|親権者|受取人|送付先)"
    r"\s*(?:氏名)?\s*[:：]?\s*[一-龥々ヶァ-ヺー・]{1,15}(?:\s+[一-龥々ヶァ-ヺー・]{1,15})?"
)
# 敬称付き: 「甲野太郎様」「ジョン・スミス様」
_HONORIFIC_RE = re.compile(r"[一-龥々ヶァ-ヺー・]{2,15}\s*(?:様|氏|殿)(?=\s|$|[、。：:,])")

# メールアドレス
_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]{1,64}@[A-Za-z0-9\-]{1,63}(?:\.[A-Za-z0-9\-]{1,63}){1,4}")

# クレジットカード番号: ラベル付き or 生の 13-19 桁連続（ハイフン/空白許容）。
# 偽陽性防止: 単なる「カード」単語単独では発火させず、番号ラベル or
# 「クレジットカード」「Visa」「MasterCard」「JCB」「AmEx」「American Express」
# 等の識別子の周辺に出るカードっぽい桁列を検出する。
_CARD_LABEL_RE = re.compile(
    r"(?:カード番号|Card\s*No\.?|Credit\s*Card|クレジット(?:カード)?|Visa|MasterCard|"
    r"JCB|AmEx|American\s*Express|Diners|ダイナース|銀聯)"
    r"[\s:：\-]*[\d０-９][\d０-９\s\-－]{12,22}[\d０-９]"
)
# ラベルなしの 13-19 桁連続（ハイフン/スペース 1 個挟み許容）。Luhn はチェックせず、
# 「数字群+区切り+数字群」の繰り返しで 13-19 桁になるパターンに限定する。
_CARD_BARE_RE = re.compile(
    r"\b[\d０-９]{4}[\-－\s]?[\d０-９]{4}[\-－\s]?[\d０-９]{4}[\-－\s]?[\d０-９]{1,7}\b"
)

# 住所: 都道府県名 + 市区町村
_PREFECTURES = (
    "北海道|青森県|岩手県|宮城県|秋田県|山形県|福島県|茨城県|栃木県|群馬県|"
    "埼玉県|千葉県|東京都|神奈川県|新潟県|富山県|石川県|福井県|山梨県|長野県|"
    "岐阜県|静岡県|愛知県|三重県|滋賀県|京都府|大阪府|兵庫県|奈良県|和歌山県|"
    "鳥取県|島根県|岡山県|広島県|山口県|徳島県|香川県|愛媛県|高知県|福岡県|"
    "佐賀県|長崎県|熊本県|大分県|宮崎県|鹿児島県|沖縄県"
)
_ADDRESS_RE = re.compile(rf"(?:{_PREFECTURES})[^\s]{{2,40}}?[市区町村郡]")

# マイナンバー: 12 桁数字（空白・ハイフン挟み許容）
_MY_NUMBER_RE = re.compile(r"\b(?:\d[\- ]?){12}\b")

# 事件番号: 「令和5年(ワ)第100号」「平成30年(家)第234号」
_CASE_NUMBER_RE = re.compile(
    r"(?:令和|平成|昭和)\s*(?:元|\d{1,2})\s*年\s*[\(（][^)）]{1,3}[\)）]\s*第\s*\d{1,6}\s*号"
)


PATTERNS: List[Tuple[str, "re.Pattern[str]", str]] = [
    ("postal_code",   _POSTAL_RE,     "郵便番号"),
    ("phone",         _PHONE_RE,      "電話番号"),
    ("email",         _EMAIL_RE,      "メールアドレス"),
    ("card_number",   _CARD_LABEL_RE, "クレジットカード番号（ラベル付き）"),
    ("card_number",   _CARD_BARE_RE,  "クレジットカード番号（13-19 桁連続）"),
    ("bank_account",  _BANK_ACCOUNT_RE, "口座番号（7桁以上）"),
    ("bank_account",  _BANK_HOLDER_RE,"口座名義"),
    ("birthdate",     _BIRTHDATE_RE,  "生年月日"),
    ("personal_name", _NAME_RE,       "当事者氏名"),
    ("personal_name", _HONORIFIC_RE,  "敬称付き氏名"),
    ("address_jp",    _ADDRESS_RE,    "住所（都道府県+市区町村）"),
    ("my_number",     _MY_NUMBER_RE,  "マイナンバー形式"),
    ("case_number",   _CASE_NUMBER_RE,"具体的事件番号"),
]

# プレースホルダ類は PII ではない（「○○様」「氏名：○○」「XXX 様」等）
_PLACEHOLDER_RE = re.compile(r"^[\s\-_○◯◎□■△▲☆★※Xx×＊\*\?？]{1,}$")


@dataclass
class Finding:
    category: str
    label: str
    excerpt: str
    cell: Optional[str]   # "B3" 等（テキストスキャン時は None）
    row: Optional[int]
    column: Optional[int]

    def to_json(self) -> Dict[str, Any]:
        d: Dict[str, Any] = {
            "category": self.category,
            "label": self.label,
            "excerpt": self.excerpt,
        }
        if self.cell:
            d["cell"] = self.cell
        if self.row is not None:
            d["row"] = self.row
        if self.column is not None:
            d["column"] = self.column
        return d


# ---------------------------------------------------------------------------
# テキストスキャン
# ---------------------------------------------------------------------------


def _excerpt_around(text: str, start: int, end: int, pad: int = 12) -> str:
    """マッチ前後に pad 文字ずつ余白を付けた抜粋。"""
    a = max(0, start - pad)
    b = min(len(text), end + pad)
    prefix = "…" if a > 0 else ""
    suffix = "…" if b < len(text) else ""
    return f"{prefix}{text[a:b]}{suffix}".replace("\n", " ")


def scan_text(text: str) -> List[Finding]:
    """1 個の文字列に対して全パターンを適用する。"""
    if not text or _PLACEHOLDER_RE.match(text):
        return []
    findings: List[Finding] = []
    seen_spans: List[Tuple[int, int, str]] = []
    for category, pattern, label in PATTERNS:
        for m in pattern.finditer(text):
            s, e = m.span()
            matched = m.group(0)
            # マッチ部分がプレースホルダのみなら除外
            if _PLACEHOLDER_RE.match(matched):
                continue
            # 同じ範囲を複数カテゴリで拾ったら最初の 1 件だけ
            overlap = any(not (e <= ss or s >= ee) for (ss, ee, _) in seen_spans)
            if overlap:
                continue
            seen_spans.append((s, e, category))
            findings.append(Finding(
                category=category,
                label=label,
                excerpt=_excerpt_around(text, s, e),
                cell=None,
                row=None,
                column=None,
            ))
    return findings


# ---------------------------------------------------------------------------
# XLSX スキャン
# ---------------------------------------------------------------------------

# DoS 防御: 個別セルがこれを超えたら先頭だけを走査し、警告を出す
MAX_CELL_LEN = 10_000
# DoS 防御: ファイルサイズがこれを超えたら警告
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB


def _col_letter(col: int) -> str:
    s = ""
    n = col
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(ord("A") + r) + s
    return s


def scan_xlsx(xlsx_path: Path, sheet: Optional[str] = None) -> Dict[str, Any]:
    """XLSX の全セル文字列を走査して PII を抽出する。

    DoS 防御:
    - セル値が `MAX_CELL_LEN` を超えたら先頭 `MAX_CELL_LEN` 文字だけ走査し、
      警告を出す（gigantic shared string をフルスキャンしない）
    - ファイルサイズが `MAX_FILE_SIZE` を超えたら警告（開く自体は通すが、
      呼出側が早期中止できるように報告する）
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError("openpyxl が必要。`pip install openpyxl` で入れてほしい。")

    warnings: List[str] = []
    truncated_cells = 0

    try:
        fsize = xlsx_path.stat().st_size
    except OSError:
        fsize = 0
    if fsize > MAX_FILE_SIZE:
        warnings.append(
            f"ファイルサイズが {fsize:,} バイトあり、スキャンに時間がかかる可能性がある "
            f"(閾値: {MAX_FILE_SIZE:,})。"
        )

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheets_to_scan = [sheet] if sheet else wb.sheetnames
    all_findings: List[Finding] = []
    for sh_name in sheets_to_scan:
        if sh_name not in wb.sheetnames:
            continue
        ws = wb[sh_name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if not isinstance(v, str):
                    s = str(v)
                else:
                    s = v
                if s.startswith("="):
                    continue
                if len(s) > MAX_CELL_LEN:
                    truncated_cells += 1
                    s = s[:MAX_CELL_LEN]
                for f in scan_text(s):
                    f.row = cell.row
                    f.column = cell.column
                    f.cell = f"{_col_letter(cell.column)}{cell.row}"
                    f.excerpt = f"[シート:{sh_name}] {f.excerpt}"
                    all_findings.append(f)

    if truncated_cells:
        warnings.append(
            f"{truncated_cells} 件のセルが {MAX_CELL_LEN:,} 文字を超えていたため "
            "先頭のみをスキャンした。長大セルに PII が埋め込まれている場合は見逃す可能性がある。"
        )

    # カテゴリ別の件数サマリー
    by_cat: Dict[str, int] = {}
    for f in all_findings:
        by_cat[f.category] = by_cat.get(f.category, 0) + 1

    return {
        "xlsx": str(xlsx_path),
        "scanned_sheets": sheets_to_scan,
        "findings": [f.to_json() for f in all_findings],
        "count": len(all_findings),
        "by_category": by_cat,
        "verdict": "clean" if not all_findings else "suspicious",
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _cmd_scan(args: argparse.Namespace) -> int:
    p = Path(args.xlsx).expanduser()
    if not p.exists():
        print(json.dumps({"error": f"xlsx が存在しない: {p}"}, ensure_ascii=False), file=sys.stderr)
        return 1
    result = scan_xlsx(p, sheet=args.sheet)
    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        if result["verdict"] == "clean":
            print(f"OK: {p.name} に PII らしき記述は検出されなかった。")
            return 0
        print(f"⚠ {p.name} に PII らしき記述が {result['count']} 件検出された:\n")
        for f in result["findings"][:50]:
            cell = f.get("cell", "—")
            print(f"  [{f['category']}] {cell}: {f['excerpt']}")
        if result["count"] > 50:
            print(f"  ... 他 {result['count'] - 50} 件")
        print()
        print("カテゴリ別:")
        for cat, n in sorted(result["by_category"].items(), key=lambda x: -x[1]):
            print(f"  {cat}: {n} 件")
    return 0


def _cmd_check(args: argparse.Namespace) -> int:
    findings = scan_text(args.text)
    print(json.dumps(
        [f.to_json() for f in findings], ensure_ascii=False, indent=2,
    ))
    return 0


def _self_test() -> int:
    """pattern の conservative 設計を検証する。"""
    ok = 0
    fail = 0

    def check(name: str, cond: bool, detail: str = "") -> None:
        nonlocal ok, fail
        status = "PASS" if cond else "FAIL"
        print(f"  [{status}] {name}{(' — ' + detail) if detail else ''}")
        if cond:
            ok += 1
        else:
            fail += 1

    cases: List[Tuple[str, str, str]] = [
        # (text, expected_category or "none", description)
        ("〒100-0001 東京都千代田区千代田1-1", "postal_code", "postal code detected"),
        ("電話: 03-1234-5678", "phone", "landline phone"),
        ("090-1234-5678", "phone", "mobile phone"),
        ("口座番号 1234567", "bank_account", "bank account with label"),
        ("口座: 123", "none", "short account number not flagged"),
        ("生年月日: 昭和50年1月1日", "birthdate", "birthdate with label"),
        ("令和5年(ワ)第100号", "case_number", "realistic case number"),
        ("○○様", "none", "placeholder name ignored"),
        ("甲野太郎様", "personal_name", "real name with honorific"),
        ("原告 山田花子", "personal_name", "party name"),
        ("東京都千代田区丸の内1丁目", "address_jp", "address with prefecture"),
        ("氏名: XXX", "none", "XXX placeholder ignored"),
        ("業務日時", "none", "generic word not flagged"),
        ("", "none", "empty string"),
        ("金額: 100,000円", "none", "money amount not flagged as PII"),
        ("番号: ___", "none", "underline placeholder ignored"),
        # --- new coverage (review feedback) ---
        ("ジョン・スミス様", "personal_name", "katakana foreign name with middle dot"),
        ("連絡先: yamada@example.co.jp", "email", "email address detected"),
        ("abc@localhost", "none", "email without TLD not flagged"),
        ("ＦＡＸ: ０３−１２３４−５６７８", "phone", "fullwidth fax number"),
        ("電話: ０９０-１２３４-５６７８", "phone", "fullwidth mobile number"),
        ("〒１００−０００１", "postal_code", "fullwidth postal code"),
        ("カード番号: 4111-1111-1111-1111", "card_number", "labeled card number"),
        ("4111 1111 1111 1111", "card_number", "bare 16-digit card number"),
        ("口座名義: 甲野太郎", "bank_account", "account holder name"),
        ("名義人: 山田花子", "bank_account", "holder label variant"),
        ("店番号 001", "none", "3-digit branch code not flagged"),
    ]

    for text, expected, desc in cases:
        got = scan_text(text)
        if expected == "none":
            check(desc, not got, f"text={text!r} got={[f.category for f in got]}")
        else:
            categories = {f.category for f in got}
            check(desc, expected in categories, f"text={text!r} got={categories}")

    print(f"\npii_scan self-test: {ok}/{ok + fail} passed")
    return 0 if fail == 0 else 1


def main() -> int:
    ap = argparse.ArgumentParser(description="XLSX/テキストの簡易 PII プリフライト")
    ap.add_argument("--self-test", action="store_true")
    sub = ap.add_subparsers(dest="command")

    p_scan = sub.add_parser("scan", help="XLSX の全セルを走査")
    p_scan.add_argument("--xlsx", required=True, help="スキャン対象 XLSX のパス")
    p_scan.add_argument("--sheet", help="対象シート名（省略時は全シート）")
    p_scan.add_argument("--json", action="store_true", help="JSON 出力")
    p_scan.set_defaults(func=_cmd_scan)

    p_check = sub.add_parser("check", help="1 文字列を判定（テスト用）")
    p_check.add_argument("text")
    p_check.set_defaults(func=_cmd_check)

    args = ap.parse_args()
    if args.self_test:
        return _self_test()
    if args.command is None:
        ap.print_help()
        return 1
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
