#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""サンプル差分ベースのテンプレートフィールド検出（v3.2.0〜）。

`/template-create` の補助。空のテンプレート XLSX と、同じレイアウトで記入済みの
サンプル XLSX を比較し、以下を **決定論的に** 抽出する:

- 単一セルフィールド（blank が空、sample に literal 値が入っているセル）
- テーブルフィールド（連続する複数行で同一カラム構造の差分領域）

LLM による「空セルの左を見て推測」ではなく、セル単位の実差分から field を特定するため
セクション見出し・小計セル・2 列ラベルなどに強い。

## 前提
- blank と sample は同じシート構造・同じ列幅・同じ結合セルを持つこと
- sample は **1 件の完全記入例** であることを推奨（複数例混在は推定誤りの温床）
- sample のセルが formula なら除外（数式セルは入力欄ではない）

## ラベル推論
各入力セル (r, c) について、以下の順で近傍の非空・非数式・非差分セルを探してラベルとする:
  1. 同じ行の左側（c-1, c-2, ... まで探索。最大 c-5 まで）
  2. 上方向（r-1, r-2, ... まで探索。最大 r-5 まで）
  3. 見つからなければ label = None（呼出側で手動ラベリング）

## タイプ推論
sample の値を観察して:
- ISO 日付・元号日付形式 → `date`
- 数値（カンマ除去後に float 化可能） → `number`
- その他 → `text`

## テーブル検出
差分セルを行・列方向にグループ化し、以下の条件に合致するクラスタを `table` にまとめる:
- 同じ列セット (c1..c2) が連続 2 行以上で差分
- headerRow = クラスタ最上行の直上の「非差分の非空行」
- 各列のラベルは headerRow の該当列の値

## CLI
```
python3 skills/_lib/template_detect.py diff --blank <blank.xlsx> --sample <sample.xlsx> [--sheet <name>]
```
JSON を stdout に出す。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl  # type: ignore
except ImportError:
    print("エラー: openpyxl が必要。`pip install openpyxl` で入れてほしい。", file=sys.stderr)
    sys.exit(2)


# ---------------------------------------------------------------------------
# 型推論
# ---------------------------------------------------------------------------

_DATE_PATTERNS = [
    re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$"),
    re.compile(r"^\d{4}年\d{1,2}月\d{1,2}日$"),
    re.compile(r"^(令和|平成|昭和|大正|明治)\s*\d{1,2}年\d{1,2}月\d{1,2}日$"),
    re.compile(r"^(R|H|S|T|M)\s*\d{1,2}[./]\d{1,2}[./]\d{1,2}$"),
]

_NUMBER_RE = re.compile(r"^-?[\d,]+(\.\d+)?$")


def _infer_type(value: Any) -> str:
    if value is None:
        return "text"
    # openpyxl は date を datetime 型で返す
    import datetime as _dt
    if isinstance(value, (_dt.date, _dt.datetime)):
        return "date"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "number"
    s = str(value).strip()
    if not s:
        return "text"
    for pat in _DATE_PATTERNS:
        if pat.match(s):
            return "date"
    # 全角数字はまず半角化してから判定
    s_half = s.translate(str.maketrans("０１２３４５６７８９、，．", "0123456789,,."))
    if _NUMBER_RE.match(s_half.replace(",", "")):
        return "number"
    return "text"


# ---------------------------------------------------------------------------
# データ構造
# ---------------------------------------------------------------------------


@dataclass
class Position:
    row: int
    column: int


@dataclass
class DetectedField:
    id: str
    label: Optional[str]
    type: str
    required: bool
    position: Optional[Position] = None
    range: Optional[Dict[str, int]] = None
    columns: Optional[List[Dict[str, Any]]] = None
    sample_value: Optional[str] = None

    def to_json(self) -> Dict[str, Any]:
        d: Dict[str, Any] = {
            "id": self.id,
            "label": self.label,
            "type": self.type,
            "required": self.required,
        }
        if self.position is not None:
            d["position"] = asdict(self.position)
        if self.range is not None:
            d["range"] = self.range
        if self.columns is not None:
            d["columns"] = self.columns
        if self.sample_value is not None:
            d["example_value"] = self.sample_value
        return d


# ---------------------------------------------------------------------------
# 読取
# ---------------------------------------------------------------------------


def _load_sheet(path: Path, sheet: Optional[str]):
    wb = openpyxl.load_workbook(path, data_only=False)
    if sheet:
        if sheet not in wb.sheetnames:
            raise ValueError(f"シート '{sheet}' が {path.name} に存在しない。候補: {wb.sheetnames}")
        ws = wb[sheet]
    else:
        ws = wb.active
    return wb, ws


def _cell_display(cell) -> Optional[str]:
    """セル値を str で返す。formula セルは None。"""
    if cell is None:
        return None
    v = cell.value
    if v is None:
        return None
    # formula セル（`=SUM(A1:A10)` 等）は入力欄ではないため除外
    if isinstance(v, str) and v.startswith("="):
        return None
    return str(v)


def _is_formula(cell) -> bool:
    if cell is None or cell.value is None:
        return False
    return isinstance(cell.value, str) and cell.value.startswith("=")


# ---------------------------------------------------------------------------
# 差分検出
# ---------------------------------------------------------------------------


def _diff_cells(blank_ws, sample_ws) -> List[Tuple[int, int, Any]]:
    """blank と sample の差分セル [(row, col, sample_value), ...] を返す。

    - blank が空 かつ sample が非空・非数式 → 入力セル候補
    - blank と sample が異なる（片方だけ埋まっていても検出）
    - 両方に同じ値 → ラベル/ヘッダ（差分ではない）
    - どちらかが数式 → 除外
    """
    # 走査範囲は両シートの max 行・列のうち大きい方
    max_row = max(blank_ws.max_row or 0, sample_ws.max_row or 0)
    max_col = max(blank_ws.max_column or 0, sample_ws.max_column or 0)

    diffs: List[Tuple[int, int, Any]] = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            b_cell = blank_ws.cell(row=r, column=c)
            s_cell = sample_ws.cell(row=r, column=c)
            if _is_formula(b_cell) or _is_formula(s_cell):
                continue
            b_val = _cell_display(b_cell)
            s_val = _cell_display(s_cell)
            if b_val == s_val:
                continue
            # blank が空 & sample が非空 → 典型的な入力欄
            # もしくは両方が値だが違う → サンプル固有（既定文言の差し替え）
            if s_val is None or s_val == "":
                # sample も空（blank だけ埋まってる）→ 無視
                continue
            diffs.append((r, c, s_cell.value))
    return diffs


# ---------------------------------------------------------------------------
# ラベル推論
# ---------------------------------------------------------------------------


def _find_label(blank_ws, r: int, c: int, diff_set: set, max_steps: int = 5) -> Optional[str]:
    """(r, c) の入力セルに対するラベルを近傍から拾う。

    探索順:
      1. 同じ行の左（c-1 → c-5）
      2. 上方向（r-1 → r-5）

    blank 上で非空・非数式・差分セットに含まれていない、かつ短文（40 文字以内）を採用。
    """
    # 左方向
    for dc in range(1, max_steps + 1):
        cc = c - dc
        if cc < 1:
            break
        if (r, cc) in diff_set:
            continue
        cell = blank_ws.cell(row=r, column=cc)
        if _is_formula(cell):
            continue
        val = _cell_display(cell)
        if val and len(val.strip()) <= 40:
            return val.strip()
    # 上方向
    for dr in range(1, max_steps + 1):
        rr = r - dr
        if rr < 1:
            break
        if (rr, c) in diff_set:
            continue
        cell = blank_ws.cell(row=rr, column=c)
        if _is_formula(cell):
            continue
        val = _cell_display(cell)
        if val and len(val.strip()) <= 40:
            return val.strip()
    return None


# ---------------------------------------------------------------------------
# テーブル検出
# ---------------------------------------------------------------------------


def _group_tables(
    diffs: List[Tuple[int, int, Any]],
    blank_ws,
) -> Tuple[List[Dict[str, Any]], List[Tuple[int, int, Any]]]:
    """差分セルのうち、同一列セットが連続 2 行以上に渡るクラスタを table にまとめる。

    戻り値: (tables, remaining_singles)
      tables = [{range, columns, rows: [(r, c, v), ...]}]
      remaining_singles = table に属さなかった差分セル
    """
    # 行ごとに列セットを集める
    by_row: Dict[int, List[Tuple[int, Any]]] = {}
    for (r, c, v) in diffs:
        by_row.setdefault(r, []).append((c, v))

    tables: List[Dict[str, Any]] = []
    consumed: set = set()
    rows_sorted = sorted(by_row.keys())

    i = 0
    while i < len(rows_sorted):
        r0 = rows_sorted[i]
        cols0 = tuple(sorted(c for c, _ in by_row[r0]))
        if len(cols0) < 2:
            i += 1
            continue
        # 連続する行で **完全に同じ列セット** がどこまで続くか
        # 以前は subset で寛容に受容していたが、それだと narrow な小計行や
        # 隣接する別テーブルを吸収してしまう。厳密一致にする。
        j = i + 1
        cluster_rows = [r0]
        while j < len(rows_sorted):
            rj = rows_sorted[j]
            if rj != cluster_rows[-1] + 1:
                break
            cols_j = tuple(sorted(c for c, _ in by_row[rj]))
            if cols_j != cols0:
                break
            cluster_rows.append(rj)
            j += 1

        if len(cluster_rows) >= 2:
            # table 候補。headerRow 推定: クラスタ最上行の 1 行上で、
            # cols0 の各列に非空ラベルがある行。
            header_row = None
            for hr in range(cluster_rows[0] - 1, max(0, cluster_rows[0] - 6), -1):
                if hr < 1:
                    break
                header_labels = []
                all_filled = True
                for c in cols0:
                    cell = blank_ws.cell(row=hr, column=c)
                    v = _cell_display(cell)
                    if not v:
                        all_filled = False
                        break
                    header_labels.append(v.strip())
                if all_filled:
                    header_row = hr
                    break

            columns = []
            for idx, c in enumerate(cols0):
                # 各列のタイプは cluster_rows の最初のセルから推定
                first_row = cluster_rows[0]
                sample_val = next((v for (rr, cc, v) in diffs if rr == first_row and cc == c), None)
                col_type = _infer_type(sample_val)
                label = None
                if header_row is not None:
                    hv = _cell_display(blank_ws.cell(row=header_row, column=c))
                    if hv:
                        label = hv.strip()
                columns.append({
                    "id": f"col_{idx + 1}",
                    "label": label,
                    "type": col_type,
                })

            tables.append({
                "range": {
                    "headerRow": header_row,
                    "dataStartRow": cluster_rows[0],
                    "startColumn": cols0[0],
                    "endRow": cluster_rows[-1],
                    "endColumn": cols0[-1],
                },
                "columns": columns,
                "rows": [(r, c, v) for r in cluster_rows for (c, v) in by_row.get(r, [])],
            })
            for r in cluster_rows:
                for (c, _) in by_row[r]:
                    consumed.add((r, c))
            i = j
        else:
            i += 1

    remaining = [(r, c, v) for (r, c, v) in diffs if (r, c) not in consumed]
    return tables, remaining


# ---------------------------------------------------------------------------
# メイン検出
# ---------------------------------------------------------------------------


def detect_fields(blank_path: Path, sample_path: Path, sheet: Optional[str] = None) -> Dict[str, Any]:
    """blank と sample を突き合わせてフィールドを検出する。

    戻り値:
    {
      "blank_file": "...",
      "sample_file": "...",
      "sheet": "...",
      "fields": [DetectedField, ...],
      "warnings": [str, ...],
    }
    """
    warnings: List[str] = []
    _, blank_ws = _load_sheet(blank_path, sheet)
    _, sample_ws = _load_sheet(sample_path, sheet)

    # sample が blank より下に広がるのは正常（データ行が増えるため）。
    # 以下のケースのみ警告する:
    #   - 列数が異なる（レイアウトがずれている可能性）
    #   - blank のほうが sample より広い（sample が不完全な可能性）
    if blank_ws.max_column != sample_ws.max_column:
        warnings.append(
            f"列数が異なる (blank: {blank_ws.max_column} 列 / sample: {sample_ws.max_column} 列)。"
            "同じレイアウトで保存してほしい。"
        )
    if (blank_ws.max_row or 0) > (sample_ws.max_row or 0):
        warnings.append(
            f"blank の行数 ({blank_ws.max_row}) が sample ({sample_ws.max_row}) より多い。"
            "sample は記入済みの完全例か確認してほしい。"
        )

    diffs = _diff_cells(blank_ws, sample_ws)
    if not diffs:
        return {
            "blank_file": str(blank_path),
            "sample_file": str(sample_path),
            "sheet": blank_ws.title,
            "fields": [],
            "warnings": warnings + ["差分セルが 0 件。sample が正しく記入済みか確認してほしい。"],
        }

    # テーブル検出
    tables, singles = _group_tables(diffs, blank_ws)

    diff_set = {(r, c) for (r, c, _) in diffs}
    detected: List[DetectedField] = []

    # 単一セルフィールド
    for idx, (r, c, v) in enumerate(singles):
        label = _find_label(blank_ws, r, c, diff_set)
        col_type = _infer_type(v)
        detected.append(DetectedField(
            id=f"field_{idx + 1}",
            label=label,
            type=col_type,
            required=True,
            position=Position(row=r, column=c),
            sample_value=str(v) if v is not None else None,
        ))

    # テーブルフィールド
    for t_idx, t in enumerate(tables):
        # テーブル全体のラベル推定: headerRow 上方 1 行
        table_label = None
        hr = t["range"]["headerRow"]
        if hr and hr >= 2:
            prev = _cell_display(blank_ws.cell(row=hr - 1, column=t["range"]["startColumn"]))
            if prev:
                table_label = prev.strip()
        detected.append(DetectedField(
            id=f"table_{t_idx + 1}",
            label=table_label,
            type="table",
            required=True,
            range=t["range"],
            columns=t["columns"],
        ))

    return {
        "blank_file": str(blank_path),
        "sample_file": str(sample_path),
        "sheet": blank_ws.title,
        "fields": [f.to_json() for f in detected],
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _cmd_diff(args: argparse.Namespace) -> int:
    blank = Path(args.blank).expanduser()
    sample = Path(args.sample).expanduser()
    if not blank.exists():
        print(json.dumps({"error": f"blank が存在しない: {blank}"}, ensure_ascii=False), file=sys.stderr)
        return 1
    if not sample.exists():
        print(json.dumps({"error": f"sample が存在しない: {sample}"}, ensure_ascii=False), file=sys.stderr)
        return 1
    try:
        result = detect_fields(blank, sample, sheet=args.sheet)
    except ValueError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        return 1
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


def _self_test() -> int:
    """stdlib + openpyxl の self-test。一時 XLSX を作って差分検出を走らせる。"""
    import tempfile
    ok = 0
    fail = 0

    def check(name: str, cond: bool, detail: str = "") -> None:
        nonlocal ok, fail
        status = "PASS" if cond else "FAIL"
        print(f"  [{status}] {name}{(' — ' + detail) if detail else ''}")
        if cond:
            ok += 1
        else:
            fail += 1

    with tempfile.TemporaryDirectory() as td:
        tdp = Path(td)
        blank_p = tdp / "blank.xlsx"
        sample_p = tdp / "sample.xlsx"

        # シナリオ 1: 単一セル + テーブル
        # blank:
        #   A1: "債権者一覧表"
        #   A3: "事件番号:"  B3: ""
        #   A5: "№"  B5: "債権者名"  C5: "金額"
        #   A6..A8: ""  B6..B8: ""  C6..C8: ""
        # sample:
        #   A3 同じ,  B3: "令和5年(ワ)第100号"
        #   A6-A8: 1/2/3  B6-B8: 山田/佐藤/鈴木  C6-C8: 100000/250000/50000

        wb_b = openpyxl.Workbook()
        ws_b = wb_b.active
        ws_b.title = "Sheet1"
        ws_b["A1"] = "債権者一覧表"
        ws_b["A3"] = "事件番号:"
        ws_b["A5"] = "№"
        ws_b["B5"] = "債権者名"
        ws_b["C5"] = "金額"
        wb_b.save(blank_p)

        wb_s = openpyxl.Workbook()
        ws_s = wb_s.active
        ws_s.title = "Sheet1"
        ws_s["A1"] = "債権者一覧表"
        ws_s["A3"] = "事件番号:"
        ws_s["B3"] = "令和5年(ワ)第100号"
        ws_s["A5"] = "№"
        ws_s["B5"] = "債権者名"
        ws_s["C5"] = "金額"
        for i, (num, name, amt) in enumerate([(1, "山田", 100000), (2, "佐藤", 250000), (3, "鈴木", 50000)]):
            ws_s.cell(row=6 + i, column=1, value=num)
            ws_s.cell(row=6 + i, column=2, value=name)
            ws_s.cell(row=6 + i, column=3, value=amt)
        wb_s.save(sample_p)

        result = detect_fields(blank_p, sample_p)
        fields = result["fields"]

        # 1. 単一セル（B3）が検出される
        single = [f for f in fields if f.get("position") and f["position"]["row"] == 3 and f["position"]["column"] == 2]
        check("1. single-cell field at B3 detected", len(single) == 1)
        check(
            "2. B3 label inferred = '事件番号:'",
            single and single[0].get("label") == "事件番号:",
            f"label={single[0].get('label') if single else None}",
        )

        # 3. テーブル検出
        tables = [f for f in fields if f["type"] == "table"]
        check("3. table detected", len(tables) == 1)
        if tables:
            t = tables[0]
            check("4. table headerRow = 5", t["range"]["headerRow"] == 5, f"got {t['range']['headerRow']}")
            check("5. table dataStartRow = 6", t["range"]["dataStartRow"] == 6)
            check("6. table endRow = 8", t["range"]["endRow"] == 8)
            col_labels = [c["label"] for c in t["columns"]]
            check("7. table column labels", col_labels == ["№", "債権者名", "金額"], f"got {col_labels}")
            col_types = [c["type"] for c in t["columns"]]
            check("8. table column types (number/text/number)", col_types == ["number", "text", "number"], f"got {col_types}")

        # 9. warnings 空
        check("9. no warnings", not result["warnings"])

        # 10. 空 sample → warning
        blank2 = tdp / "blank2.xlsx"
        sample2 = tdp / "sample2.xlsx"
        wb2 = openpyxl.Workbook()
        wb2.active["A1"] = "空のテンプレ"
        wb2.save(blank2)
        wb3 = openpyxl.Workbook()
        wb3.active["A1"] = "空のテンプレ"  # 全く同じ = 差分なし
        wb3.save(sample2)
        result2 = detect_fields(blank2, sample2)
        check("10. empty diff yields warning", bool(result2["warnings"]))

        # 11. テーブル検出の厳密一致: 3 列テーブルの後に 2 列の小計行が続くケース。
        # subset 条件では小計行を同じテーブルに吸収してしまうが、exact 一致では別扱い。
        blank3 = tdp / "blank3.xlsx"
        sample3 = tdp / "sample3.xlsx"
        wb_b3 = openpyxl.Workbook()
        ws_b3 = wb_b3.active
        ws_b3["A1"] = "№"; ws_b3["B1"] = "名称"; ws_b3["C1"] = "金額"
        ws_b3["A5"] = "合計:"
        wb_b3.save(blank3)
        wb_s3 = openpyxl.Workbook()
        ws_s3 = wb_s3.active
        ws_s3["A1"] = "№"; ws_s3["B1"] = "名称"; ws_s3["C1"] = "金額"
        ws_s3["A2"] = 1; ws_s3["B2"] = "aaa"; ws_s3["C2"] = 100
        ws_s3["A3"] = 2; ws_s3["B3"] = "bbb"; ws_s3["C3"] = 200
        ws_s3["A4"] = 3; ws_s3["B4"] = "ccc"; ws_s3["C4"] = 300
        ws_s3["A5"] = "合計:"; ws_s3["C5"] = 600  # 小計: A と C だけ埋まっている（2 列）
        wb_s3.save(sample3)
        result3 = detect_fields(blank3, sample3)
        tbls3 = [f for f in result3["fields"] if f["type"] == "table"]
        check(
            "11. narrower subtotal row NOT absorbed into table",
            len(tbls3) == 1 and tbls3[0]["range"]["endRow"] == 4,
            f"tables={tbls3}",
        )

    print(f"\ntemplate_detect self-test: {ok}/{ok + fail} passed")
    return 0 if fail == 0 else 1


def main() -> int:
    ap = argparse.ArgumentParser(description="サンプル差分ベースのテンプレートフィールド検出")
    ap.add_argument("--self-test", action="store_true")
    sub = ap.add_subparsers(dest="command")

    p_diff = sub.add_parser("diff", help="blank と sample の差分からフィールドを検出")
    p_diff.add_argument("--blank", required=True, help="空のテンプレート XLSX のパス")
    p_diff.add_argument("--sample", required=True, help="記入済みサンプル XLSX のパス")
    p_diff.add_argument("--sheet", help="対象シート名（省略時は active sheet）")
    p_diff.set_defaults(func=_cmd_diff)

    args = ap.parse_args()
    if args.self_test:
        return _self_test()
    if args.command is None:
        ap.print_help()
        return 1
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
