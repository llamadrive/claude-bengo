#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""bengo-toolkit 同梱テンプレートのレジストリと `/template-install` バックエンド。

同梱テンプレートはプラグインディレクトリの `templates/_bundled/{id}/` に
`{id}.yaml` + `{id}.xlsx` のペアで配置される。`_registry.yaml` が
利用可能なテンプレートのメタデータを列挙する。

`/template-install` は現在の workspace の `templates/` ディレクトリ、または
ユーザーの `~/.claude-bengo/templates/` へ同梱テンプレートをコピーする。
case スコープでは workspace 未初期化でも現在のフォルダを自動初期化して続行する。

## スコープ名（v3.5.0 リネーム）

v3.5.0 で `global` → `user` にリネームした。従来の「事務所グローバル」は
実際には端末別・lawyer 別なので user と呼ぶのが正確。次回リリースで本格的な
firm スコープ（Shared Drive 同期フォルダ）を追加する予定。

旧名 `global` は CLI / 関数引数で一代だけ受け付ける（使用時に stderr 警告）。
次回リリースで削除する。

## CLI

```
python3 skills/_lib/template_lib.py list [--format json]
    同梱テンプレート一覧を表示する
python3 skills/_lib/template_lib.py install <id> [--scope case|user] [--replace]
    指定テンプレートを case または user スコープへコピーする
python3 skills/_lib/template_lib.py show <id>
    テンプレートのメタデータ詳細を表示する
```
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# プラグインディレクトリ配下の同梱ディレクトリ
PLUGIN_ROOT = Path(__file__).resolve().parent.parent.parent
BUNDLED_DIR = PLUGIN_ROOT / "templates" / "_bundled"
REGISTRY_FILE = BUNDLED_DIR / "_registry.yaml"
MANIFEST_FILE = BUNDLED_DIR / "_manifest.sha256"

# 同梱テンプレート ID は conservative regex（ファイルシステム安全）
BUNDLED_ID_RE = re.compile(r"^[a-z0-9][-a-z0-9_]{0,63}$")


# ---------------------------------------------------------------------------
# スコープ正規化（v3.5.0）
# ---------------------------------------------------------------------------
# 旧名 `global` を入力として受け付け、内部では `user` に正規化する。
# 使用時は stderr に非推奨警告を出す。次回リリースで `global` は削除。

_SCOPE_DEPRECATION_WARNED = False


def _normalize_scope(scope: str, *, context: str = "") -> str:
    """入力 scope を正規化する。

    - "case" → "case"
    - "firm" → "firm"
    - "user" → "user"
    - "global" → "user"（非推奨警告付き）
    - それ以外 → ValueError

    Args:
        scope: 入力スコープ名
        context: 警告メッセージに含める呼出コンテキスト（例: "--scope", "save_user_template"）
    """
    global _SCOPE_DEPRECATION_WARNED
    if scope == "case":
        return "case"
    if scope == "firm":
        return "firm"
    if scope == "user":
        return "user"
    if scope == "global":
        if not _SCOPE_DEPRECATION_WARNED:
            _SCOPE_DEPRECATION_WARNED = True
            ctx = f" ({context})" if context else ""
            print(
                f"注意: scope \"global\" は v3.5.0 で \"user\" にリネームされた{ctx}。"
                "次回リリースで \"global\" は削除される。\"user\" に書き換えてほしい。",
                file=sys.stderr,
            )
        return "user"
    raise ValueError(f"scope は 'case' / 'firm' / 'user' のいずれか。指定値: {scope!r}")


class FirmUnavailableError(RuntimeError):
    """firm スコープが unconfigured または unreachable のときに投げる。

    呼出側はユーザーに `/template-firm-setup <path>` の案内を出すべき。
    `state` フィールドに `"unconfigured"` または `"unreachable"` が入る。
    """

    def __init__(self, state: str, path: Optional[str]):
        self.state = state
        self.path = path
        if state == "unconfigured":
            msg = (
                "firm スコープが設定されていない。"
                "`/template-firm-setup <path>` で OS 同期クライアントの "
                "shared folder を 1 度設定してほしい。"
            )
        else:
            msg = (
                f"firm スコープのパス ({path}) に到達できない。"
                "Drive for desktop / Dropbox 等の同期クライアントが起動しているか、"
                "フォルダが削除/移動されていないか確認してほしい。"
            )
        super().__init__(msg)


def _resolve_scope_dir(scope: str, *, ensure: bool = True) -> "Path":
    """scope → 書込/読込先ディレクトリ。

    case: ensure_workspace → templates_dir
    firm: firm_templates_dir（reachable 必須）。FirmUnavailableError を投げうる
    user: ensure_user_templates_dir

    `ensure=False` の場合、case でも `ensure_workspace` を呼ばずに既存パスのみ
    返す（読込専用パスとして使う）。
    """
    ws = _load_workspace()
    if scope == "user":
        return ws.ensure_user_templates_dir()
    if scope == "firm":
        fs = ws.firm_status()
        if fs["state"] != "reachable":
            raise FirmUnavailableError(fs["state"], fs.get("path"))
        # firm dir は admin が作るもの。プラグインからは mkdir しない。
        return Path(fs["path"])
    # case
    if ensure:
        ws.ensure_workspace()
    d = ws.templates_dir()
    if ensure:
        d.mkdir(parents=True, exist_ok=True)
    return d


# ---------------------------------------------------------------------------
# workspace.py の遅延ロード（v3.0.0〜）
# ---------------------------------------------------------------------------


def _load_workspace():
    import importlib
    here = str(Path(__file__).resolve().parent)
    added = False
    if here not in sys.path:
        sys.path.insert(0, here)
        added = True
    try:
        return importlib.import_module("workspace")
    finally:
        if added:
            try:
                sys.path.remove(here)
            except ValueError:
                pass


# ---------------------------------------------------------------------------
# レジストリ読込
# ---------------------------------------------------------------------------


def _parse_registry_entry(block: str) -> Optional[Dict[str, str]]:
    """`_registry.yaml` のエントリブロックを簡易パースする。

    矢印マーカー（`- id: ...`）で始まるブロックを dict に変換する。
    複数行値は `|` スカラー相当で扱う（旧 metadata リーダと同じ規約）。
    """
    entry: Dict[str, str] = {}
    current_key: Optional[str] = None
    current_lines: List[str] = []

    def flush() -> None:
        nonlocal current_key, current_lines
        if current_key is not None:
            entry[current_key] = "\n".join(current_lines).rstrip()
        current_key = None
        current_lines = []

    for raw in block.splitlines():
        stripped = raw.rstrip()
        if current_key is not None and raw.startswith("    "):
            current_lines.append(raw[4:])
            continue
        flush()
        # 先頭 `- id:` を通常の key: value として扱う
        s = stripped.lstrip("- ")
        if not s or s.startswith("#"):
            continue
        if ":" not in s:
            continue
        k, _, v = s.partition(":")
        k = k.strip()
        v = v.strip()
        if v == "|":
            current_key = k
            current_lines = []
            continue
        if len(v) >= 2 and v[0] == v[-1] and v[0] in ('"', "'"):
            v = v[1:-1]
        entry[k] = v
    flush()
    return entry if entry else None


def load_registry() -> List[Dict[str, str]]:
    """`_registry.yaml` を読み、エントリのリストを返す。

    簡易パーサ。行頭 `- id:` を各エントリの開始と見なし、次の `- id:` か
    EOF までをそのエントリに含める。インデントは矢印マーカー直下の
    4 スペース前提。
    """
    if not REGISTRY_FILE.exists():
        return []
    text = REGISTRY_FILE.read_text(encoding="utf-8")

    # `templates:` 配下の `- id: ...` エントリを抽出する。
    in_templates = False
    blocks: List[List[str]] = []
    current: List[str] = []

    for raw in text.splitlines():
        s = raw.rstrip()
        if not in_templates:
            if s.startswith("templates:"):
                in_templates = True
            continue
        # 次の top-level key で終了
        if s and not s.startswith(" ") and not s.startswith("#") and ":" in s and not s.lstrip().startswith("-"):
            break
        stripped = s.lstrip()
        if stripped.startswith("- id:"):
            if current:
                blocks.append(current)
            current = [s]
        elif current:
            current.append(s)

    if current:
        blocks.append(current)

    entries: List[Dict[str, str]] = []
    for b in blocks:
        # 各ブロック: 先頭行から `  ` プレフィックスを剥がす
        normalized = []
        for line in b:
            if line.startswith("    "):
                normalized.append(line[2:])  # インデントを 2 段剥がす
            elif line.startswith("  - "):
                normalized.append(line[2:])
            elif line.startswith("  "):
                normalized.append(line[2:])
            else:
                normalized.append(line)
        entry = _parse_registry_entry("\n".join(normalized))
        if entry and "id" in entry:
            # 同梱ファイルの実在確認
            bid = entry["id"]
            yaml_p = BUNDLED_DIR / bid / f"{bid}.yaml"
            xlsx_p = BUNDLED_DIR / bid / f"{bid}.xlsx"
            entry["_yaml_exists"] = str(yaml_p.exists())
            entry["_xlsx_exists"] = str(xlsx_p.exists())
            entry["_yaml_path"] = str(yaml_p)
            entry["_xlsx_path"] = str(xlsx_p)
            entries.append(entry)

    return entries


def find_template(bundled_id: str) -> Optional[Dict[str, str]]:
    """指定 ID の同梱テンプレートを返す。無ければ None。"""
    if not BUNDLED_ID_RE.match(bundled_id):
        return None
    for e in load_registry():
        if e.get("id") == bundled_id:
            return e
    return None


# ---------------------------------------------------------------------------
# マニフェスト検証（v2.6.1〜）
#
# プラグインクローン後に悪意ある第三者が templates/_bundled/ 配下の
# YAML/XLSX を書き換える「テンプレートすり替え攻撃」を検知する。
# マニフェストは `scripts/build_bundled_forms.py` により生成され、
# リポジトリに git 追跡される。install 時に対象ファイルの SHA-256 を
# 計算し、マニフェスト記載のハッシュと照合する。
# ---------------------------------------------------------------------------


def _load_manifest() -> Dict[str, str]:
    """`_manifest.sha256` を読んで {相対パス: sha256} の dict を返す。

    マニフェストが存在しない場合は空 dict（後方互換。install は警告付きで続行）。
    """
    if not MANIFEST_FILE.exists():
        return {}
    result: Dict[str, str] = {}
    try:
        for line in MANIFEST_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(None, 1)
            if len(parts) == 2:
                result[parts[1]] = parts[0]
    except OSError:
        return {}
    return result


def _sha256(path: Path) -> str:
    """ファイルの SHA-256（16進）を返す。"""
    import hashlib
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _verify_bundled_integrity(bundled_id: str) -> Optional[str]:
    """指定テンプレートのファイルがマニフェストと一致するか検証する。

    戻り値:
        None: 整合（またはマニフェスト未整備で検証スキップ）
        str : 不一致の詳細メッセージ
    """
    manifest = _load_manifest()
    if not manifest:
        # F-037: マニフェストが存在しない場合は install を拒否する。
        # 旧実装は "warn + proceed" だったが、これはマニフェスト削除を攻撃経路
        # として使える。明示的に --skip-integrity フラグでのみ許可する。
        return (
            "_manifest.sha256 が存在しない。テンプレート整合性検証ができないため install を中止する。"
            "プラグインクローンが破損している可能性がある。`/plugin install bengo-toolkit@bengo-toolkit` で再取得するか、"
            "やむをえず先に進める場合は --skip-integrity を明示指定してほしい。"
        )

    for ext in ("yaml", "xlsx"):
        rel = f"{bundled_id}/{bundled_id}.{ext}"
        expected = manifest.get(rel)
        if not expected:
            return f"マニフェストに {rel} のエントリがない（同梱ファイルが不完全）"
        full = BUNDLED_DIR / rel
        if not full.exists():
            return f"{rel} が存在しない"
        actual = _sha256(full)
        if actual != expected:
            return (
                f"テンプレート '{bundled_id}' のファイル {rel} が改ざんされている可能性がある。\n"
                f"  expected: {expected}\n"
                f"  actual  : {actual}\n"
                f"プラグインを `/plugin install bengo-toolkit@bengo-toolkit` で再取得することを検討してほしい。"
            )
    return None


# ---------------------------------------------------------------------------
# インストール
# ---------------------------------------------------------------------------


def install_template(
    bundled_id: str,
    replace: bool = False,
    skip_integrity: bool = False,
    scope: str = "case",
) -> Dict[str, str]:
    """同梱テンプレートを指定スコープのテンプレートディレクトリへコピーする。

    v3.3.0 でデフォルトを `case` に戻した。lawyer が無意識に user-wide へ配置
    してしまい、他案件の shadowing が混線するリスクがあるため。端末全体で
    使い回したい場合は `scope="user"` を明示、あるいは登録後に `/template-promote`
    を使う。

    - `scope="user"` → `~/.claude-bengo/templates/{id}.{yaml,xlsx}`
      端末別・lawyer 別に全案件で共有する書式。workspace 初期化不要。
    - `scope="case"` → `<workspace>/.claude-bengo/templates/{id}.{yaml,xlsx}`
      現在の案件フォルダに限定。workspace が未初期化なら silently 初期化する。

    v3.5.0: `scope="global"` は `"user"` にリネームされた。旧名も一代だけ受け付ける
    （stderr 警告付き、次回リリースで削除）。

    戻り値: {yaml_dst, xlsx_dst, scope, bundled_id, replaced, integrity_verified,
             workspace_root (scope=case のみ)}
    """
    scope = _normalize_scope(scope, context="install_template")

    ws = _load_workspace()

    entry = find_template(bundled_id)
    if not entry:
        raise ValueError(f"同梱テンプレート '{bundled_id}' は見つからない。/template-install で一覧を確認してほしい。")

    # 整合性検証（v2.6.1〜）
    integrity_verified = False
    if not skip_integrity:
        err = _verify_bundled_integrity(bundled_id)
        if err:
            raise ValueError(f"テンプレート整合性検証失敗: {err}")
        integrity_verified = MANIFEST_FILE.exists()

    src_yaml = Path(entry["_yaml_path"])
    src_xlsx = Path(entry["_xlsx_path"])
    if not src_yaml.exists():
        raise FileNotFoundError(f"{src_yaml} が存在しない（レジストリと同梱ファイルの不整合）")
    if not src_xlsx.exists():
        raise FileNotFoundError(f"{src_xlsx} が存在しない（レジストリと同梱ファイルの不整合）")

    dst_dir = _resolve_scope_dir(scope)
    if scope == "case":
        workspace_root: Optional[str] = str(ws.resolve_workspace())
    else:
        workspace_root = None

    dst_yaml = dst_dir / f"{bundled_id}.yaml"
    dst_xlsx = dst_dir / f"{bundled_id}.xlsx"

    if (dst_yaml.exists() or dst_xlsx.exists()) and not replace:
        scope_labels = {
            "user": "ユーザースコープ",
            "firm": "事務所スコープ",
            "case": "この案件フォルダ",
        }
        scope_label = scope_labels.get(scope, scope)
        raise FileExistsError(
            f"{scope_label}に既に '{bundled_id}' が存在する。上書きには --replace を指定してほしい。"
        )

    # NOTE: install_template は bundled template が source。manifest 検証で
    # 改ざん検知済み・PII フリーが保証されているため、追加の PII スキャンは
    # 不要。user 由来の XLSX は save_user_template / promote_template が
    # PII gate を担当する。

    shutil.copy2(src_yaml, dst_yaml, follow_symlinks=False)
    shutil.copy2(src_xlsx, dst_xlsx, follow_symlinks=False)

    result: Dict[str, str] = {
        "bundled_id": bundled_id,
        "scope": scope,
        "yaml_dst": str(dst_yaml),
        "xlsx_dst": str(dst_xlsx),
        "replaced": str(replace),
        "integrity_verified": str(integrity_verified),
    }
    if workspace_root is not None:
        result["workspace_root"] = workspace_root
    return result


# ---------------------------------------------------------------------------
# PII code-gate（v3.3.0-iter1〜）
# ---------------------------------------------------------------------------


class PIIFoundError(ValueError):
    """user スコープへの書込時に PII が検出された場合に投げる。

    発生時は **ユーザー側での overridable ではない**。findings 属性に
    検出されたレコード一覧を持つ。開発バックドア: 環境変数
    `CLAUDE_BENGO_ALLOW_PII_ON_GLOBAL=1` で続行（テスト・CI 用）。
    ユーザー向けフラグではない。env var 名は v3.5.0 リネーム前に設定された
    ため `GLOBAL` のままにしている（CI 互換性維持）。
    """
    def __init__(self, findings: List[Dict], xlsx_path: Path):
        self.findings = findings
        self.xlsx_path = xlsx_path
        preview = ", ".join(
            f"{f.get('cell','?')}[{f.get('category','?')}]"
            for f in findings[:5]
        )
        more = f" 他 {len(findings) - 5} 件" if len(findings) > 5 else ""
        super().__init__(
            f"{xlsx_path.name} に PII らしき記述が {len(findings)} 件検出された ({preview}{more})。"
            " user スコープへの保存を拒否する。case スコープで登録するか、XLSX 側で PII を削除してから再実行してほしい。"
        )


def _check_pii_for_user(xlsx_path: Path) -> None:
    """XLSX が user スコープに保存されても安全か検証する。

    PII 検出時は `PIIFoundError` を投げる（ユーザーの override 不可）。

    開発者バックドア: `CLAUDE_BENGO_ALLOW_PII_ON_GLOBAL=1` で findings を無視
    して続行する。テスト・CI 用の escape hatch で、env var を明示設定する
    必要があるため偶発的な bypass は起きにくい。env var 名は v3.5.0 リネーム
    前に設定されたため `GLOBAL` のままにしている（CI 互換性維持）。
    """
    here = Path(__file__).resolve().parent
    sys_path_added = False
    if str(here) not in sys.path:
        sys.path.insert(0, str(here))
        sys_path_added = True
    try:
        import importlib
        pii_scan = importlib.import_module("pii_scan")
    except ImportError:
        raise PIIFoundError(
            [{"cell": "N/A", "category": "scanner_unavailable"}], xlsx_path,
        )
    finally:
        if sys_path_added:
            try:
                sys.path.remove(str(here))
            except ValueError:
                pass

    result = pii_scan.scan_xlsx(xlsx_path)
    if result.get("verdict") == "clean":
        return

    if os.environ.get("CLAUDE_BENGO_ALLOW_PII_ON_GLOBAL") == "1":
        print(
            f"警告: CLAUDE_BENGO_ALLOW_PII_ON_GLOBAL=1 により "
            f"{len(result.get('findings', []))} 件の PII 検出を無視して user 保存を続行。",
            file=sys.stderr,
        )
        return
    raise PIIFoundError(result.get("findings", []), xlsx_path)


# v3.5.0 後方互換エイリアス（内部呼出用。次回リリースで削除）
def _check_pii_for_global(xlsx_path: Path) -> None:
    _check_pii_for_user(xlsx_path)


def save_user_template(
    source_xlsx: Path,
    template_id: str,
    *,
    scope: str = "case",
    yaml_content: Optional[str] = None,
    replace: bool = False,
) -> Dict[str, str]:
    """ユーザー作成テンプレートを保存する（code-level PII guard 付き）。

    `/template-create` の Write + copy_file.py の代替。SKILL は本関数を叩くことで
    user スコープ時の PII チェックを **必ず** 通過する。

    v3.5.0: `scope="global"` は `"user"` にリネーム。旧名は一代だけ受け付ける。

    Args:
      source_xlsx: 元 XLSX のパス
      template_id: 登録 ID（ファイル名、BUNDLED_ID_RE で検証済みの英数字）
      scope: "case" または "user"（旧 "global" も一代は受け付ける）
      yaml_content: 書き込む YAML 定義（None なら既存 YAML を期待）
      replace: 既存を上書きするか

    Raises:
      PIIFoundError: scope="user" かつ XLSX に PII 検出
      ValueError: ID 不正
      FileExistsError: replace=False で既存衝突
    """
    if not BUNDLED_ID_RE.match(template_id):
        raise ValueError(f"無効なテンプレート ID: {template_id!r}")
    scope = _normalize_scope(scope, context="save_user_template")

    source = Path(source_xlsx).expanduser()
    if not source.exists() or not source.is_file():
        raise FileNotFoundError(f"source xlsx が存在しない: {source}")

    # ** CRITICAL: user / firm 保存前に PII スキャン（code-enforced） **
    # case スコープは PII チェックを通さない（案件フォルダ内なので意図的に
    # クライアントデータを含むことが正常）。
    if scope in ("user", "firm"):
        _check_pii_for_user(source)

    dst_dir = _resolve_scope_dir(scope)

    dst_yaml = dst_dir / f"{template_id}.yaml"
    dst_xlsx = dst_dir / f"{template_id}.xlsx"

    if (dst_yaml.exists() or dst_xlsx.exists()) and not replace:
        raise FileExistsError(
            f"{dst_dir} に既に '{template_id}' が存在する。上書きには replace=True を指定してほしい。"
        )

    if yaml_content is not None:
        dst_yaml.write_text(yaml_content, encoding="utf-8")
    elif not dst_yaml.exists():
        raise FileNotFoundError(
            f"yaml_content 未指定で、既存 YAML ({dst_yaml}) もない。YAML 内容を渡してほしい。"
        )
    shutil.copy2(source, dst_xlsx, follow_symlinks=False)

    return {
        "id": template_id,
        "scope": scope,
        "yaml_path": str(dst_yaml),
        "xlsx_path": str(dst_xlsx),
        "pii_scanned": "True" if scope in ("user", "firm") else "False",
    }


# ---------------------------------------------------------------------------
# 昇格／降格（case ↔ global）
# ---------------------------------------------------------------------------


def _move_template(
    template_id: str,
    *,
    src_scope: str,
    dst_scope: str,
    replace: bool,
    keep_original: bool,
) -> Dict[str, str]:
    """`{id}.yaml` + `{id}.xlsx` を src_scope → dst_scope に移す共通ロジック。

    - `promote`: case → user, keep_original=False（moveに相当。案件側の shadowing を解消）
    - `demote`:  user → case, keep_original=True（copy に相当。他案件の user は維持）
    - `scope` へのコピー後、src_scope のファイルを削除するかは keep_original で制御

    v3.5.0: `global` → `user` リネーム。src_scope/dst_scope に旧名 `global` が
    渡されたら `user` に正規化して警告。

    戻り値: {id, src_scope, dst_scope, src_yaml, src_xlsx, dst_yaml, dst_xlsx,
             replaced, kept_original}
    """
    src_scope = _normalize_scope(src_scope, context="_move_template.src_scope")
    dst_scope = _normalize_scope(dst_scope, context="_move_template.dst_scope")
    if src_scope == dst_scope:
        raise ValueError(f"src と dst のスコープが同じ: {src_scope}")
    if not BUNDLED_ID_RE.match(template_id):
        raise ValueError(f"無効なテンプレート ID: {template_id!r}")

    SCOPE_LABELS = {"case": "この案件フォルダ", "firm": "事務所スコープ", "user": "ユーザースコープ"}

    def _scope_paths(scope: str) -> Tuple[Path, Path, Path]:
        d = _resolve_scope_dir(scope)
        return d, d / f"{template_id}.yaml", d / f"{template_id}.xlsx"

    _, src_yaml, src_xlsx = _scope_paths(src_scope)
    if not src_yaml.exists() or not src_xlsx.exists():
        raise FileNotFoundError(
            f"{SCOPE_LABELS[src_scope]}にテンプレート '{template_id}' が見つからない "
            f"(yaml={src_yaml.exists()}, xlsx={src_xlsx.exists()})"
        )

    _, dst_yaml, dst_xlsx = _scope_paths(dst_scope)
    dst_yaml_existed = dst_yaml.exists()
    dst_xlsx_existed = dst_xlsx.exists()
    if (dst_yaml_existed or dst_xlsx_existed) and not replace:
        raise FileExistsError(
            f"{SCOPE_LABELS[dst_scope]}に既に '{template_id}' が存在する。上書きには --replace を指定してほしい。"
        )

    # --- 原子的コピー（stage-then-rename）---
    # .tmp にコピー → 両方揃ったら元に rename で上書き（rename は POSIX で原子的）。
    # 途中で失敗したら dst を元状態に戻す。これにより:
    #   - コピー途中でクラッシュしても dst が半端な状態で残らない
    #   - --replace 時の「yaml は新、xlsx は旧」の不整合が起きない
    stage_yaml = dst_yaml.with_suffix(dst_yaml.suffix + ".staging")
    stage_xlsx = dst_xlsx.with_suffix(dst_xlsx.suffix + ".staging")
    # 既存の staging 残骸があれば掃除
    for p in (stage_yaml, stage_xlsx):
        if p.exists():
            try:
                p.unlink()
            except OSError:
                pass

    try:
        shutil.copy2(src_yaml, stage_yaml, follow_symlinks=False)
        shutil.copy2(src_xlsx, stage_xlsx, follow_symlinks=False)
    except Exception:
        for p in (stage_yaml, stage_xlsx):
            try:
                if p.exists():
                    p.unlink()
            except OSError:
                pass
        raise

    # --replace 時に dst の現状を別 backup に退避（ロールバック用）
    backup_yaml = dst_yaml.with_suffix(dst_yaml.suffix + ".backup") if dst_yaml_existed else None
    backup_xlsx = dst_xlsx.with_suffix(dst_xlsx.suffix + ".backup") if dst_xlsx_existed else None
    try:
        if backup_yaml:
            dst_yaml.replace(backup_yaml)
        if backup_xlsx:
            dst_xlsx.replace(backup_xlsx)
        stage_yaml.replace(dst_yaml)
        stage_xlsx.replace(dst_xlsx)
    except Exception:
        # ロールバック: backup があれば戻す / staging を消す
        for src, dst in ((backup_yaml, dst_yaml), (backup_xlsx, dst_xlsx)):
            if src and src.exists():
                try:
                    src.replace(dst)
                except OSError:
                    pass
        for p in (stage_yaml, stage_xlsx):
            if p.exists():
                try:
                    p.unlink()
                except OSError:
                    pass
        raise
    else:
        for p in (backup_yaml, backup_xlsx):
            if p and p.exists():
                try:
                    p.unlink()
                except OSError:
                    pass

    delete_failed = False
    delete_error: Optional[str] = None
    if not keep_original:
        # 原本削除も best-effort だが、失敗時は result に delete_failed=True を載せて
        # caller が stale な case 側を検知できるようにする。サイレントに成功扱いしない。
        for p in (src_yaml, src_xlsx):
            try:
                p.unlink()
            except OSError as e:
                delete_failed = True
                delete_error = str(e)
                print(
                    f"警告: 元ファイル削除に失敗: {p} — {e}。"
                    f"コピーは成功したが case 側に残骸があるため、"
                    f"手動で削除するか /template-promote の再実行が必要。",
                    file=sys.stderr,
                )

    result: Dict[str, str] = {
        "id": template_id,
        "src_scope": src_scope,
        "dst_scope": dst_scope,
        "src_yaml": str(src_yaml),
        "src_xlsx": str(src_xlsx),
        "dst_yaml": str(dst_yaml),
        "dst_xlsx": str(dst_xlsx),
        "replaced": str(replace),
        "kept_original": str(keep_original),
        "delete_failed": str(delete_failed),
    }
    if delete_error:
        result["delete_error"] = delete_error
    return result


def promote_template(
    template_id: str,
    replace: bool = False,
    *,
    to: str = "user",
) -> Dict[str, str]:
    """案件スコープ → ユーザー or 事務所スコープに **移動**（案件側から削除）。

    `to`:
      "user" — この端末・lawyer 全案件で見える（既定、従来どおり）
      "firm" — 事務所全員で共有（v3.6.0〜、Shared Drive 同期フォルダ）

    典型的なユースケース:
      最初は案件専用で登録した書式が他案件でも使えると気づいたとき、
      昇格して案件側の shadowing を解消する。事務所全員に展開したい場合は `to=firm`。

    **PII gate**: 昇格先が user / firm のいずれでも、移動前に pii_scan を通す
    （code-enforced hard block）。PII 検出時は `PIIFoundError` を投げて移動しない。
    """
    if not BUNDLED_ID_RE.match(template_id):
        raise ValueError(f"無効なテンプレート ID: {template_id!r}")
    to = _normalize_scope(to, context="promote_template.to")
    if to == "case":
        raise ValueError("promote 先に 'case' は指定できない（昇格にならないため）")

    ws = _load_workspace()
    case_dir = ws.templates_dir()
    src_xlsx = case_dir / f"{template_id}.xlsx"
    if src_xlsx.exists():
        _check_pii_for_user(src_xlsx)  # PIIFoundError を投げうる（user/firm 共通）

    return _move_template(
        template_id,
        src_scope="case",
        dst_scope=to,
        replace=replace,
        keep_original=False,
    )


def demote_template(
    template_id: str,
    replace: bool = False,
    *,
    from_: str = "user",
) -> Dict[str, str]:
    """ユーザー or 事務所スコープ → 案件スコープに **コピー**（src は維持）。

    `from_`:
      "user" — 端末ローカルから降格（既定、従来どおり）
      "firm" — 事務所共有から降格（v3.6.0〜）

    典型的なユースケース:
      標準書式を特定案件だけ微修正したいとき、src のコピーを案件スコープに
      置いて shadowing を作る。他案件は従来どおり src を参照するため src は残す。
    """
    from_ = _normalize_scope(from_, context="demote_template.from_")
    if from_ == "case":
        raise ValueError("demote 元に 'case' は指定できない（降格にならないため）")
    return _move_template(
        template_id,
        src_scope=from_,
        dst_scope="case",
        replace=replace,
        keep_original=True,
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _cmd_list(args: argparse.Namespace) -> int:
    entries = load_registry()
    if args.format == "json":
        print(json.dumps(entries, ensure_ascii=False, indent=2))
        return 0
    if not entries:
        print("同梱テンプレートが見つからない。")
        return 0
    print("利用可能な同梱テンプレート:")
    print()
    # カテゴリで group 化
    by_cat: Dict[str, List[Dict[str, str]]] = {}
    for e in entries:
        by_cat.setdefault(e.get("category", "その他"), []).append(e)
    for cat in sorted(by_cat):
        print(f"  [{cat}]")
        for e in by_cat[cat]:
            warn = "" if e.get("_yaml_exists") == "True" and e.get("_xlsx_exists") == "True" else " (⚠ 同梱ファイル不足)"
            title = e.get("title", e["id"])
            print(f"    {e['id']:<22}  {title}{warn}")
            if e.get("description"):
                desc = e["description"].splitlines()[0][:80]
                print(f"                              {desc}")
        print()
    print("インストール方法:")
    print("  /template-install <id>         — 現在の案件フォルダに同梱テンプレートをコピー")
    print("  /template-install <id> --replace — 既存を上書き")
    return 0


def _cmd_install(args: argparse.Namespace) -> int:
    # v3.0.0: workspace 解決（CWD walk-up）。

    try:
        result = install_template(
            args.bundled_id,
            replace=args.replace,
            skip_integrity=getattr(args, "skip_integrity", False),
            scope=getattr(args, "scope", "case"),
        )
    except FirmUnavailableError as e:
        print(json.dumps(
            {"error": str(e), "code": "firm_unavailable", "state": e.state, "path": e.path},
            ensure_ascii=False,
        ), file=sys.stderr)
        return 6
    except (ValueError, FileNotFoundError) as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        return 1
    except FileExistsError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        return 3

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


def _cmd_show(args: argparse.Namespace) -> int:
    e = find_template(args.bundled_id)
    if not e:
        print(json.dumps({"error": f"'{args.bundled_id}' が見つからない"}, ensure_ascii=False), file=sys.stderr)
        return 1
    print(json.dumps(e, ensure_ascii=False, indent=2))
    return 0


def _cmd_promote(args: argparse.Namespace) -> int:
    try:
        result = promote_template(args.id, replace=args.replace, to=args.to)
    except FirmUnavailableError as e:
        print(json.dumps(
            {"error": str(e), "code": "firm_unavailable", "state": e.state, "path": e.path},
            ensure_ascii=False,
        ), file=sys.stderr)
        return 6
    except PIIFoundError as e:
        print(
            json.dumps(
                {
                    "error": str(e),
                    "code": "pii_found",
                    "findings": e.findings[:20],
                    "total_findings": len(e.findings),
                },
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )
        return 4
    except FileNotFoundError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        return 1
    except FileExistsError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        return 3
    except ValueError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        return 1
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


def _cmd_save_user(args: argparse.Namespace) -> int:
    yaml_content = None
    if args.yaml_file:
        yaml_content = Path(args.yaml_file).expanduser().read_text(encoding="utf-8")
    try:
        result = save_user_template(
            Path(args.source),
            args.id,
            scope=args.scope,
            yaml_content=yaml_content,
            replace=args.replace,
        )
    except FirmUnavailableError as e:
        print(json.dumps(
            {"error": str(e), "code": "firm_unavailable", "state": e.state, "path": e.path},
            ensure_ascii=False,
        ), file=sys.stderr)
        return 6
    except PIIFoundError as e:
        print(
            json.dumps(
                {
                    "error": str(e),
                    "code": "pii_found",
                    "findings": e.findings[:20],
                    "total_findings": len(e.findings),
                },
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )
        return 4
    except FileExistsError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        return 3
    except (ValueError, FileNotFoundError) as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        return 1
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


def _cmd_demote(args: argparse.Namespace) -> int:
    try:
        result = demote_template(args.id, replace=args.replace, from_=args.from_)
    except FirmUnavailableError as e:
        print(json.dumps(
            {"error": str(e), "code": "firm_unavailable", "state": e.state, "path": e.path},
            ensure_ascii=False,
        ), file=sys.stderr)
        return 6
    except FileNotFoundError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        return 1
    except FileExistsError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        return 3
    except ValueError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        return 1
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


def _self_test() -> int:
    """サニティチェック。同梱テンプレートのファイル整合性 + promote/demote を検証する。"""
    entries = load_registry()
    if not entries:
        print("  [WARN] no bundled templates (registry empty)")
        return 0
    all_ok = True
    for e in entries:
        bid = e.get("id", "?")
        if e.get("_yaml_exists") != "True":
            print(f"  [FAIL] {bid}: yaml missing at {e.get('_yaml_path')}")
            all_ok = False
        elif e.get("_xlsx_exists") != "True":
            print(f"  [FAIL] {bid}: xlsx missing at {e.get('_xlsx_path')}")
            all_ok = False
        else:
            yaml_p = Path(e["_yaml_path"])
            text = yaml_p.read_text(encoding="utf-8")
            if "fields:" not in text:
                print(f"  [FAIL] {bid}: yaml missing 'fields:' block")
                all_ok = False
                continue
            print(f"  [PASS] {bid}: yaml + xlsx present, fields block ok")
    print(f"\ntemplate_lib: registry has {len(entries)} entries")

    # --- promote / demote self-test（一時 HOME + workspace で隔離実行） ---
    import os
    import tempfile
    ok = 0
    fail = 0

    def check(name: str, cond: bool, detail: str = "") -> None:
        nonlocal ok, fail
        status = "PASS" if cond else "FAIL"
        print(f"  [{status}] promote/demote: {name}{(' — ' + detail) if detail else ''}")
        if cond:
            ok += 1
        else:
            fail += 1

    ws = _load_workspace()
    with tempfile.TemporaryDirectory() as td:
        fake_home = Path(td) / "home"
        fake_home.mkdir()
        case_root = Path(td) / "case-A"
        case_root.mkdir()

        # GLOBAL_ROOT / HOME を一時ディレクトリへ差し替え
        orig_global_root = ws.GLOBAL_ROOT
        orig_global_cfg = ws.GLOBAL_CONFIG_FILE
        orig_home = os.environ.get("HOME")
        ws.GLOBAL_ROOT = fake_home / ".claude-bengo"
        ws.GLOBAL_CONFIG_FILE = ws.GLOBAL_ROOT / "global.json"
        os.environ["HOME"] = str(fake_home)
        cwd_before = Path.cwd()
        os.chdir(case_root)
        try:
            ws.ensure_workspace()
            # case 側に手動でダミーテンプレート（{id}.yaml + 実 xlsx）を置く
            import openpyxl
            case_tdir = ws.templates_dir()
            case_tdir.mkdir(parents=True, exist_ok=True)
            (case_tdir / "promo-test.yaml").write_text("id: promo-test\ntitle: 昇格テスト\nfields: []\n", encoding="utf-8")
            wb_clean = openpyxl.Workbook()
            wb_clean.active["A1"] = "氏名"  # ラベルのみ、PII なし
            wb_clean.active["B1"] = "金額"
            wb_clean.save(case_tdir / "promo-test.xlsx")

            # 1. promote: case → user（moveなので case 側が消える）
            result = promote_template("promo-test")
            user_tdir = ws.user_templates_dir()
            check(
                "1. promoted yaml exists in user",
                (user_tdir / "promo-test.yaml").exists(),
                f"dst_yaml={result['dst_yaml']}",
            )
            check(
                "2. promoted xlsx exists in user",
                (user_tdir / "promo-test.xlsx").exists(),
            )
            check(
                "3. case side removed after promote",
                not (case_tdir / "promo-test.yaml").exists() and not (case_tdir / "promo-test.xlsx").exists(),
            )
            check(
                "3b. promote result dst_scope == 'user' (v3.5.0 rename)",
                result.get("dst_scope") == "user",
                f"dst_scope={result.get('dst_scope')}",
            )

            # 4. demote: user → case（copy。user は残る）
            demo_result = demote_template("promo-test")
            check(
                "4. demoted yaml exists in case",
                (case_tdir / "promo-test.yaml").exists(),
                f"dst_yaml={demo_result['dst_yaml']}",
            )
            check(
                "5. user side preserved after demote",
                (user_tdir / "promo-test.yaml").exists(),
            )

            # 6. collision without --replace raises
            raised = False
            try:
                demote_template("promo-test")
            except FileExistsError:
                raised = True
            check("6. demote collision without --replace raises", raised)

            # 7. --replace succeeds
            demo_result2 = demote_template("promo-test", replace=True)
            check("7. demote with --replace succeeds", demo_result2["replaced"] == "True")

            # 8. unknown id raises FileNotFoundError
            raised2 = False
            try:
                promote_template("does-not-exist")
            except FileNotFoundError:
                raised2 = True
            check("8. promote of unknown id raises FileNotFoundError", raised2)

            # 9. invalid id regex raises ValueError
            raised3 = False
            try:
                promote_template("../etc/passwd")
            except ValueError:
                raised3 = True
            check("9. promote of path-traversal id raises ValueError", raised3)

            # 10. --replace atomic swap: failure during stage → original retained
            # sample: put a valid template in user, then try to demote with --replace
            # and verify backup cleanup (no .backup leaked). Success path smoke test.
            _ = demote_template("promo-test", replace=True)
            leftovers = list(case_tdir.glob("*.backup")) + list(case_tdir.glob("*.staging"))
            leftovers += list(user_tdir.glob("*.backup")) + list(user_tdir.glob("*.staging"))
            check("10. no .staging/.backup left after --replace", not leftovers, f"leftovers={leftovers}")

            # 10a. backward compat: _move_template accepts legacy "global" arg
            # (wrapped via _normalize_scope; emits deprecation warning)
            # Re-put template in case for the compat test
            (case_tdir / "compat-test.yaml").write_text("id: compat-test\nfields: []\n", encoding="utf-8")
            wb_compat = openpyxl.Workbook()
            wb_compat.active["A1"] = "ラベル"
            wb_compat.save(case_tdir / "compat-test.xlsx")
            compat_result = _move_template(
                "compat-test",
                src_scope="case",
                dst_scope="global",   # legacy name — should be accepted, normalized to "user"
                replace=False,
                keep_original=False,
            )
            check(
                "10a. legacy dst_scope='global' is accepted and normalized to 'user'",
                compat_result.get("dst_scope") == "user",
                f"dst_scope={compat_result.get('dst_scope')}",
            )

            # 10b. promote with PII in xlsx must raise PIIFoundError (code-enforced)
            pii_case_dir = case_tdir  # already the case templates dir
            pii_id = "pii-test"
            (pii_case_dir / f"{pii_id}.yaml").write_text(
                f"id: {pii_id}\ntitle: PIIテスト\nfields: []\n", encoding="utf-8",
            )
            wb = openpyxl.Workbook()
            wb.active["A1"] = "氏名: 甲野太郎様"  # PII 混入
            wb.active["A2"] = "東京都千代田区千代田1-1"
            wb.save(pii_case_dir / f"{pii_id}.xlsx")

            raised_pii = False
            try:
                promote_template(pii_id)
            except PIIFoundError:
                raised_pii = True
            check("10b. promote with PII raises PIIFoundError", raised_pii)
            # case side must still be there (promote aborted)
            check(
                "10c. case xlsx preserved after failed promote",
                (pii_case_dir / f"{pii_id}.xlsx").exists(),
            )

            # 10d. ALLOW_PII_ON_GLOBAL=1 bypasses (test / CI escape hatch)
            os.environ["CLAUDE_BENGO_ALLOW_PII_ON_GLOBAL"] = "1"
            try:
                r_override = promote_template(pii_id)
                check("10d. ALLOW_PII_ON_GLOBAL=1 bypasses", r_override["dst_scope"] == "user")
            finally:
                os.environ.pop("CLAUDE_BENGO_ALLOW_PII_ON_GLOBAL", None)

            # 10e. save_user_template to user with PII is refused
            pii_src = Path(td) / "user-template.xlsx"
            wb2 = openpyxl.Workbook()
            wb2.active["A1"] = "原告 山田花子"
            wb2.save(pii_src)
            raised2 = False
            try:
                save_user_template(
                    pii_src, "user-pii-test", scope="user",
                    yaml_content="id: user-pii-test\nfields: []\n",
                )
            except PIIFoundError:
                raised2 = True
            check("10e. save_user_template(scope=user) with PII raises", raised2)

            # 10f. case scope ignores PII (local is OK)
            r_case = save_user_template(
                pii_src, "user-pii-test", scope="case",
                yaml_content="id: user-pii-test\nfields: []\n",
            )
            check("10f. save_user_template(scope=case) with PII succeeds", r_case["scope"] == "case")

            # 10g. save_user_template accepts legacy scope="global"
            raised_compat = False
            try:
                save_user_template(
                    pii_src, "user-pii-test", scope="global",   # legacy
                    yaml_content="id: user-pii-test\nfields: []\n",
                )
            except PIIFoundError:
                raised_compat = True   # Still blocked by PII, confirming the scope was accepted
            check("10g. save_user_template accepts legacy scope='global'", raised_compat)

            # 11. delete_failed flag surfaces when unlink fails (simulate by making
            # src dir read-only on POSIX. Skipped if not root and chmod can't deny).
            # Instead, pre-open a file handle and try on Windows? On macOS unlink
            # succeeds even with read-only parent dir if user owns it. Fall back:
            # check the result shape includes delete_failed key.
            # Re-promote to test the key.
            (case_tdir / "delete-test.yaml").write_text("id: delete-test\nfields: []\n", encoding="utf-8")
            wb_del = openpyxl.Workbook()
            wb_del.active["A1"] = "金額"
            wb_del.save(case_tdir / "delete-test.xlsx")
            r = promote_template("delete-test")
            check("11. promote result includes delete_failed key", "delete_failed" in r)
            check("12. delete_failed=False on normal promote", r.get("delete_failed") == "False")

            # ---- 13. firm scope (PR-1 of firm-scope phase) ----------------
            # 13a. promote --to firm without firm-setup raises FirmUnavailableError
            (case_tdir / "firm-promo.yaml").write_text("id: firm-promo\nfields: []\n", encoding="utf-8")
            wb_fp = openpyxl.Workbook()
            wb_fp.active["A1"] = "ラベル"
            wb_fp.save(case_tdir / "firm-promo.xlsx")
            ws.unset_firm_templates_path()
            raised_fu = False
            try:
                promote_template("firm-promo", to="firm")
            except FirmUnavailableError as e:
                raised_fu = True
                check("13a. FirmUnavailableError state='unconfigured'", e.state == "unconfigured")
            check("13a-pre. promote --to firm raises when unconfigured", raised_fu)

            # 13b. set firm path → promote --to firm succeeds, case side removed
            firm_dir_t = Path(td) / "firm-shared-tlib"
            firm_dir_t.mkdir(parents=True, exist_ok=True)
            ws.set_firm_templates_path(firm_dir_t)
            r_firm = promote_template("firm-promo", to="firm")
            check(
                "13b. promote --to firm dst_scope='firm'",
                r_firm.get("dst_scope") == "firm",
                f"dst_scope={r_firm.get('dst_scope')}",
            )
            check(
                "13c. firm xlsx exists in firm dir",
                (firm_dir_t / "firm-promo.xlsx").exists(),
            )
            check(
                "13d. case side removed after promote --to firm",
                not (case_tdir / "firm-promo.xlsx").exists(),
            )

            # 13e. demote --from firm copies back to case (firm preserved)
            r_dem = demote_template("firm-promo", from_="firm")
            check(
                "13e. demote --from firm dst_scope='case'",
                r_dem.get("dst_scope") == "case",
                f"dst_scope={r_dem.get('dst_scope')}",
            )
            check(
                "13f. firm side preserved after demote",
                (firm_dir_t / "firm-promo.xlsx").exists(),
            )

            # 13g. promote --to firm with PII raises PIIFoundError (firm gate same as user)
            (case_tdir / "firm-pii.yaml").write_text("id: firm-pii\nfields: []\n", encoding="utf-8")
            wb_fpii = openpyxl.Workbook()
            wb_fpii.active["A1"] = "氏名: 甲野太郎様"
            wb_fpii.active["A2"] = "東京都千代田区千代田1-1"
            wb_fpii.save(case_tdir / "firm-pii.xlsx")
            raised_pii_firm = False
            try:
                promote_template("firm-pii", to="firm")
            except PIIFoundError:
                raised_pii_firm = True
            check("13g. promote --to firm blocks PII (same gate as user)", raised_pii_firm)
            check(
                "13h. case xlsx preserved after PII-blocked firm promote",
                (case_tdir / "firm-pii.xlsx").exists(),
            )

            # 13i. save_user_template scope=firm with clean xlsx succeeds
            clean_src = Path(td) / "firm-save-clean.xlsx"
            wb_cs = openpyxl.Workbook()
            wb_cs.active["A1"] = "金額"  # ラベルのみ
            wb_cs.save(clean_src)
            r_save = save_user_template(
                clean_src, "firm-save-test", scope="firm",
                yaml_content="id: firm-save-test\nfields: []\n",
            )
            check(
                "13i. save_user_template(scope=firm) succeeds for clean xlsx",
                r_save.get("scope") == "firm" and r_save.get("pii_scanned") == "True",
                f"r_save={r_save}",
            )
            check(
                "13j. firm save_user_template wrote files to firm dir",
                (firm_dir_t / "firm-save-test.xlsx").exists(),
            )

            # 13k. install_template --scope firm runs PII gate (no easy way to test
            # rejection without bundled PII fixture; just verify install succeeds for
            # a clean bundled template — exercising the firm path)
            # Pick the first available bundled template and install it to firm.
            entries_for_firm = [e for e in load_registry() if e.get("_yaml_exists") == "True"][:1]
            if entries_for_firm:
                bundled_id_firm = entries_for_firm[0]["id"]
                r_inst_firm = install_template(bundled_id_firm, scope="firm")
                check(
                    "13k. install_template(scope=firm) succeeds for clean bundled template",
                    r_inst_firm.get("scope") == "firm",
                    f"r_inst_firm={r_inst_firm}",
                )

            ws.unset_firm_templates_path()

            # ---- 13l-13o. CLI _cmd_install / _cmd_save_user catch FirmUnavailableError ----
            # Both code paths must return exit 6 with structured error, not crash.
            # We exercise them by calling the function objects directly with a forged
            # argparse namespace, since spawning a subprocess would re-init env/HOME.
            class _NS:
                def __init__(self, **kw): self.__dict__.update(kw)

            # 13l. _cmd_install with --scope firm while unconfigured
            import io as _io_test
            saved_err = sys.stderr
            sys.stderr = _io_test.StringIO()
            try:
                rc_inst = _cmd_install(_NS(
                    bundled_id="creditor-list",
                    replace=False,
                    skip_integrity=False,
                    scope="firm",
                ))
            finally:
                err_inst = sys.stderr.getvalue()
                sys.stderr = saved_err
            check(
                "13l. _cmd_install firm-unconfigured returns exit 6",
                rc_inst == 6,
                f"rc={rc_inst}",
            )
            check(
                "13m. _cmd_install firm-unconfigured emits structured error JSON",
                '"code": "firm_unavailable"' in err_inst and '"state": "unconfigured"' in err_inst,
                f"stderr={err_inst[:200]}",
            )

            # 13n. _cmd_save_user with --scope firm while unconfigured
            clean_src_cli = Path(td) / "cli-firm-clean.xlsx"
            wb_cli = openpyxl.Workbook()
            wb_cli.active["A1"] = "金額"
            wb_cli.save(clean_src_cli)
            yaml_cli = Path(td) / "cli-firm-clean.yaml"
            yaml_cli.write_text("id: cli-firm-clean\nfields: []\n", encoding="utf-8")

            saved_err = sys.stderr
            sys.stderr = _io_test.StringIO()
            try:
                rc_save = _cmd_save_user(_NS(
                    source=str(clean_src_cli),
                    id="cli-firm-clean",
                    scope="firm",
                    yaml_file=str(yaml_cli),
                    replace=False,
                ))
            finally:
                err_save = sys.stderr.getvalue()
                sys.stderr = saved_err
            check(
                "13n. _cmd_save_user firm-unconfigured returns exit 6",
                rc_save == 6,
                f"rc={rc_save}",
            )
            check(
                "13o. _cmd_save_user firm-unconfigured emits structured error JSON",
                '"code": "firm_unavailable"' in err_save and '"state": "unconfigured"' in err_save,
                f"stderr={err_save[:200]}",
            )
        finally:
            ws.GLOBAL_ROOT = orig_global_root
            ws.GLOBAL_CONFIG_FILE = orig_global_cfg
            if orig_home is not None:
                os.environ["HOME"] = orig_home
            else:
                os.environ.pop("HOME", None)
            os.chdir(cwd_before)

    print(f"\npromote/demote self-test: {ok}/{ok + fail} passed")
    return 0 if (all_ok and fail == 0) else 1


def main() -> int:
    ap = argparse.ArgumentParser(description="claude-bengo 同梱テンプレート管理")
    ap.add_argument("--self-test", action="store_true", help="同梱テンプレートの整合性をチェックする")
    sub = ap.add_subparsers(dest="command")

    p_list = sub.add_parser("list", help="同梱テンプレートの一覧")
    p_list.add_argument("--format", choices=["text", "json"], default="text")

    p_show = sub.add_parser("show", help="テンプレートの詳細")
    p_show.add_argument("bundled_id")

    p_inst = sub.add_parser("install", help="現在の案件フォルダまたは事務所全体にインストール")
    p_inst.add_argument("bundled_id")
    p_inst.add_argument("--replace", action="store_true", help="既存を上書き")
    p_inst.add_argument(
        "--scope",
        choices=("case", "firm", "user", "global"),
        default="case",
        help="case=この案件のみ（既定） / firm=事務所全員で共有（要 /template-firm-setup） / "
             "user=この端末・lawyer 全案件 / global=旧名（user の deprecated alias）",
    )
    p_inst.add_argument(
        "--skip-integrity",
        action="store_true",
        help="マニフェスト検証をスキップ（非推奨、デバッグ用）",
    )

    p_prom = sub.add_parser(
        "promote",
        help="案件スコープ → user / firm へ移動（PII 検出時は拒否）",
    )
    p_prom.add_argument("id")
    p_prom.add_argument("--replace", action="store_true", help="dst 側に既存があれば上書き")
    p_prom.add_argument(
        "--to",
        choices=("user", "firm"),
        default="user",
        help="昇格先スコープ（既定 user）",
    )

    p_save = sub.add_parser(
        "save-user",
        help="ユーザー作成テンプレートを指定スコープに保存（user/firm は PII code-gate 付き）",
    )
    p_save.add_argument("--source", required=True, help="ソース XLSX パス")
    p_save.add_argument("--id", required=True, help="テンプレート ID")
    p_save.add_argument("--scope", choices=("case", "firm", "user", "global"), default="case")
    p_save.add_argument("--yaml-file", help="書き込む YAML 定義のファイル（省略時は既存を期待）")
    p_save.add_argument("--replace", action="store_true", help="既存を上書き")

    p_dem = sub.add_parser(
        "demote",
        help="user / firm → 案件スコープへコピー（src は維持）",
    )
    p_dem.add_argument("id")
    p_dem.add_argument("--replace", action="store_true", help="case 側に既存があれば上書き")
    p_dem.add_argument(
        "--from",
        dest="from_",
        choices=("user", "firm"),
        default="user",
        help="降格元スコープ（既定 user）",
    )

    args = ap.parse_args()

    if args.self_test:
        return _self_test()

    if args.command is None:
        ap.print_help()
        return 1

    handlers = {
        "list": _cmd_list,
        "install": _cmd_install,
        "show": _cmd_show,
        "promote": _cmd_promote,
        "demote": _cmd_demote,
        "save-user": _cmd_save_user,
    }
    return handlers[args.command](args)


if __name__ == "__main__":
    sys.exit(main())
